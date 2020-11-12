import logging
import warnings
from itertools import islice

import numpy as np
import openpyxl
import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
from pymoo.model.repair import Repair

logger = logging.getLogger(__name__)


class KDORepair(Repair):
    """Implements the Knowledge-driven Design Optimization (KDO) framework by providing a common standard of
    taking user inputs and performing online innovization."""
    def __init__(self, n_var, var_cluster=None, user_input_file_name='user_input.xlsx'):
        self.n_var = n_var
        # A relation matrix, R(x, f, g) which defines pairwise relationships among variables (x), objectives (f) and
        # constraints (g). This will be used by the online innovization process and also as a user interface.
        self.relation_matrix = {}
        # The probability with which an offspring will be repaired using the user-provided relation. This value is
        # updated regularly based on the performance, and is never allowed to go below a minimum value. The user
        # will be able to see the current confidence level through a color coding of cell (i, j), where red represents
        # the last user input relation at (i, j) had a very low survival rate, and green for good survival rate.
        self.relation_confidence = {}

        # Relation matrix suggested by innovization. Only the elements corresponding to the filled location of the
        # relation matrix are filled by innovization suggestions. For example, for a relation at (i, j) like
        # "I,<=,0.5", the innovization matrix will have a suggestion for the user, something like "I,>,0.8". This
        # denotes that the innovization thinks a relation of xi > xj is correct with 80% confidence. User has the option
        # of accepting the input or overriding it.
        self.relation_matrix_innovization = {}
        # Snapshot of the last user input relation matrix (currently unused). To be used if online innov and user input
        # operating simultaneously.
        self.last_user_input = {}

        # The excel file to be used for user interaction
        self.user_input_file_path = user_input_file_name

        # If k clusters of variables are defined, the relation matrix will be split into k submatrices, one for each
        # cluster. Each cluster defines a class of variables that are likely to possess relationships among
        # themselves. This will also make it easier for the user to provide input. var_cluster will be defined as dicts
        # if provided, so the keys will act as the cluster name and an indicator of what the variable represents.
        self.var_cluster = var_cluster

        self.DEFAULT_CLUSTER_KEY = 'relation'

        if self.var_cluster is None:
            # If no clusters specified, the (n_var x n_var) relation matrix will consist of all possible pairwise
            # relations.
            self.var_cluster = {self.DEFAULT_CLUSTER_KEY: np.arange(self.n_var)}
            self.relation_matrix[self.DEFAULT_CLUSTER_KEY] = pd.DataFrame(data=[[None for _ in range(n_var)] for _ in range(n_var)],
                                                            columns=[str(i) for i in range(n_var)])
            self.relation_confidence[self.DEFAULT_CLUSTER_KEY] = np.zeros([n_var, n_var])

            self.relation_matrix_innovization[self.DEFAULT_CLUSTER_KEY] = pd.DataFrame(
                data=[[None for _ in range(n_var)] for _ in range(n_var)],
                columns=[str(i) for i in range(n_var)])
            self.last_user_input[self.DEFAULT_CLUSTER_KEY] = pd.DataFrame(
                data=[[None for _ in range(n_var)] for _ in range(n_var)],
                columns=[str(i) for i in range(n_var)])
        else:
            for key in self.var_cluster.keys():
                # Create a submatrix corresponding to each cluster.
                submatrix_size = len(self.var_cluster[key])
                self.relation_matrix[key] = pd.DataFrame(
                    data=[[None for _ in range(submatrix_size)] for _ in range(submatrix_size)],
                    columns=[str(i) for i in range(submatrix_size)])
                self.relation_confidence[key] = np.zeros([submatrix_size, submatrix_size])

                self.relation_matrix_innovization[key] = pd.DataFrame(
                    data=[[None for _ in range(submatrix_size)] for _ in range(submatrix_size)],
                    columns=[str(i) for i in range(submatrix_size)])
                self.last_user_input[key] = pd.DataFrame(
                    data=[[None for _ in range(submatrix_size)] for _ in range(submatrix_size)],
                    columns=[str(i) for i in range(submatrix_size)])

        self.alpha = 0.5
        self.min_probability = 0.1
        self.repair_interval = 50  # Generation gap between repairs

        self.relation_dictionary = {'Inequality': 'I',
                                    'Greater Than': '>', 'Less Than': '<',
                                    'Greater Than or Equal To': '>=', 'Less Than or Equal To': '<=',
                                    'Equality': 'E', 'Equal To': '=',
                                    'Inequality Innovization': '<>',
                                    'Power Law': 'P',
                                    'Normal Distribution': 'N', 'Uniform Distribution': 'U', 'Constant': 'C'}

    def find_cluster_by_index(self, var_index):
        """Finds which cluster a variable x(i) belongs to."""
        for key in self.var_cluster.keys():
            if var_index in self.var_cluster[key]:
                return key

        return None

    def write_relation_matrix_to_excel(self):
        """Write the relation matrix to an excel output file. Users can modify this to provide own input."""
        def rgb_to_hex(rgb):
            return '%02x%02x%02x' % rgb
        r_rgb = [int((255 * n)) for n in np.arange(0, 1.01, 0.1)]
        g_rgb = [int(255 * (1 - n)) for n in np.arange(0, 1.01, 0.1)]
        b_rgb = 0
        r_rgb.reverse()
        g_rgb.reverse()

        # Create a workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        for key in self.relation_matrix.keys():
            # Each variable cluster has its own separate worksheet.
            ws = wb.create_sheet(title=key)

            # Convert relation matrix data frame to excel sheet
            relation_df_to_row = dataframe_to_rows(self.relation_matrix[key], index=True, header=True)
            innov_df_to_row = dataframe_to_rows(self.relation_matrix_innovization[key], index=True, header=True)
            for indx, (r, r_innov) in enumerate(zip(relation_df_to_row, innov_df_to_row)):
                # KLUGE: Prevents extra blank row from being written below the column headers
                if len(r) == 1 and r[0] is None:
                    continue
                ws.append(r)
                if indx > 0:
                    for i, val in enumerate(r):
                        if i == 0 or r[i] is None:
                            continue
                        dv = DataValidation(type="list", formula1=f'"{val}","{r_innov[i]}"', allow_blank=True)
                        dv.showErrorMessage = False
                        ws.add_data_validation(dv)
                        dv.add(ws.cell(indx, i + 1))

                        confidence = self.relation_confidence[key][indx - 2, i - 1]
                        color_indx = int(np.round(confidence * 10))
                        rgb_color = rgb_to_hex((r_rgb[color_indx], g_rgb[color_indx], b_rgb)).upper()
                        # print(confidence, color_indx, rgb_color)
                        # ws.cell(indx, i + 1).font = Font(color=rgb_to_hex((r_rgb[color_indx], g_rgb[color_indx], b_rgb)))
                        ws.cell(indx, i + 1).fill = PatternFill("solid", fgColor=rgb_color)

            # Highlight row and column headers
            for cell in ws['A'] + ws[1]:
                cell.style = 'Pandas'

        wb.save(self.user_input_file_path)

    def validate_input(self, input_relation_matrix):
        """Validate the input entered by the user and remove invalid inputs."""
        return input_relation_matrix

    def read_user_input_excel(self):
        """Read user input from excel file."""
        wb_input = openpyxl.load_workbook(self.user_input_file_path)
        for key in self.relation_matrix.keys():
            ws_input = wb_input[key]

            data = ws_input.values
            cols = next(data)[1:]
            data = list(data)
            idx = [r[0] for r in data]
            data = (islice(r, 1, None) for r in data)
            input_relation_matrix = pd.DataFrame(data, index=idx, columns=cols)
            self.relation_matrix[key] = self.validate_input(input_relation_matrix)
            self.last_user_input[key] = self.relation_matrix[key].copy()
            self.relation_confidence[key] = np.nan_to_num(
                self.relation_matrix[key].applymap(lambda x: self.parse_relation_str(x, 2)), copy=True, nan=0.0, posinf=None, neginf=None)

    def parse_relation_str(self, rel_str, return_flag=None):
        if rel_str is None or type(rel_str) is not str:
            return -1

        split_str = rel_str.split(',')
        rel_type = split_str[0]  # Type of relation
        rel_params = split_str[1:-1]  # Relation parameters (example, for normal distribution it has <mean, std dev.>
        rel_confidence = float(split_str[-1])  # Relation confidence (between 0 to 1)

        if return_flag is None:
            return rel_type, rel_params, rel_confidence
        elif return_flag == 0:
            return rel_type
        elif return_flag == 1:
            return rel_params
        elif return_flag == 2:
            return rel_confidence
        else:
            print("Invalid return flag. Returning fully-split relationship string.")
            return rel_type, rel_params, rel_confidence

    def get_user_feedback(self):
        """Present existing results to user and gain feedback."""
        input_complete_flag = False
        while not input_complete_flag:
            user_response = input(f'Program paused. Please modify input spreadsheet. Type y when done.\n')
            if user_response == 'y' or user_response == 'Y':
                input_complete_flag = True

        self.read_user_input_excel()

    def update_probability(self, user_input_survival_rate):
        """Update the probability of each relation being applied."""
        for key in self.relation_confidence.keys():
            self.relation_confidence[key] = np.maximum(self.min_probability,
                                                       (self.alpha * user_input_survival_rate[key])
                                                       + ((1 - self.alpha) * self.relation_confidence[key]))

    def get_rule_conformity_data(self, x):
        """Find out how many population members in x follow the user-specified relation matrix."""
        pass

    def _do(self, problem, x, **kwargs):
        # if kwargs['algorithm'].n_gen % self.repair_interval != 0:
        #     return x

        # logger.info(f"Performing user input repair. gen {kwargs['algorithm'].n_gen}")

        x_repaired = np.copy(x)  # Repaired offspring population
        # Element (i,j) = 1 indicates that x(i,j) was repaired. Potentially useful for studying the interdependencies
        # of multiple rules
        var_repaired = np.zeros(x_repaired.shape)

        for pop_indx in range(x.shape[0]):
            # For every offspring iterate over relations present in each row of relation matrix
            for key in self.relation_matrix.keys():
                relation_matrix = self.relation_matrix[key]
                # Get the positions of the decision variable in the current cluster
                var_index = self.var_cluster[key]
                # Variable limits. Used for repair operations
                xl = kwargs['algorithm'].problem.xl
                xu = kwargs['algorithm'].problem.xu

                # Perform repair row-wise
                for row_index, row in relation_matrix.iterrows():
                    r = np.random.rand()
                    # TODO: Check duplicate relations on super and subdiagonal matrix elements
                    for col_index, col in enumerate(list(relation_matrix)):
                        # FIXME: Only superdiagonal elements are considered now.
                        if row_index > col_index:
                            continue
                        print(row_index, col_index, row[col])
                        # When col_index = row_index, indicates relations involving one variable
                        if row[col_index] is None:
                            continue
                        relation = self.parse_relation_str(row[col_index])
                        rel_type = relation[0]
                        rel_params = relation[1]
                        rel_confidence = relation[2]  # Relation confidence (between 0 to 1)

                        if col_index == row_index:
                            if rel_type == self.relation_dictionary['Normal Distribution']:
                                # Relation parameters for normal distribution has format <mean, std dev.>
                                mean, std = float(rel_params[0]), float(rel_params[1])
                                if (mean - std) <= x_repaired[pop_indx, var_index[col_index]] <= (mean + std):
                                    # Ignore if x already within 1 std of mean
                                    continue
                                elif r <= rel_confidence:
                                    x_repaired[pop_indx, var_index[col_index]] = np.random.normal(loc=mean, scale=std)
                                    var_repaired[pop_indx, var_index[col_index]] = 1
                            elif rel_type == self.relation_dictionary['Uniform Distribution']:
                                # Relation parameters for uniform distribution has format <lower limit, upper limit>
                                ll, ul = float(rel_params[0]), float(rel_params[1])
                                if ll <= x_repaired[pop_indx, var_index[col_index]] <= ul:
                                    # Ignore if x already within range
                                    continue
                                elif r <= rel_confidence:
                                    x_repaired[pop_indx, var_index[col_index]] = np.random.uniform(ll, ul)
                                    var_repaired[pop_indx, var_index[col_index]] = 1
                            elif rel_type == self.relation_dictionary['Constant']:
                                # If user wants variable to be set to a constant value
                                c = float(rel_params[0])
                                if x_repaired[pop_indx, var_index[col_index]] == c:
                                    continue
                                elif r <= rel_confidence:
                                    x_repaired[pop_indx, var_index[col_index]] = c
                                    var_repaired[pop_indx, var_index[col_index]] = 1
                            else:
                                warnings.warn(f"Relation type {rel_type} specified at ({row_index},{col_index}) of key {key} not "
                                              f"recognized. Skipping.")
                        else:
                            # Pairwise variable relation
                            # TODO: Separate >= and > cases.
                            if rel_type == self.relation_dictionary['Inequality']:
                                if rel_params[0] == self.relation_dictionary['Greater Than'] or rel_params[0] == self.relation_dictionary['Greater Than or Equal To']:
                                    if x_repaired[pop_indx, var_index[row_index]] > x_repaired[pop_indx, var_index[col_index]]:
                                        # Ignore if x already satisfies inequality
                                        continue
                                    elif r <= rel_confidence:
                                        if x_repaired[pop_indx, var_index[row_index]] < xl[var_index[col_index]]:
                                            warnings.warn(f"Greater than relation requested. But x{var_index[row_index]} < xl{var_index[col_index]}. Skipping.")
                                            continue
                                        else:
                                            x_repaired[pop_indx, var_index[col_index]] = \
                                                xl[var_index[col_index]] + np.random.rand() * (x_repaired[pop_indx, var_index[row_index]] - xl[var_index[col_index]])
                                            var_repaired[pop_indx, var_index[col_index]] = 1
                                elif rel_params[0] == self.relation_dictionary['Less Than'] or rel_params[0] == self.relation_dictionary['Less Than or Equal To']:
                                    if x_repaired[pop_indx, var_index[row_index]] < x_repaired[pop_indx, var_index[col_index]]:
                                        # Ignore if x already satisfies inequality
                                        continue
                                    elif r <= rel_confidence:
                                        if x_repaired[pop_indx, var_index[row_index]] > xu[var_index[col_index]]:
                                            warnings.warn(f"Greater than relation requested. But x{var_index[row_index]} > xu{var_index[col_index]}. Skipping.")
                                            continue
                                        else:
                                            x_repaired[pop_indx, var_index[col_index]] = \
                                                x_repaired[pop_indx, var_index[row_index]] + np.random.rand() * (xu[var_index[col_index]] - x_repaired[pop_indx, var_index[row_index]])
                                            var_repaired[pop_indx, var_index[col_index]] = 1
                                # elif rel_params[0] == self.relation_dictionary['Greater Than or Equal To']:
                                #     pass
                                # elif rel_params[0] == self.relation_dictionary['Less Than or Equal To']:
                                #     pass
                                elif rel_params[0] == self.relation_dictionary['Inequality Innovization']:
                                    pass
                                else:
                                    warnings.warn(f"Inequality relation specified at ({row_index},{col_index}) of key {key} not "
                                                  f"recognized. Skipping.")
                            elif rel_type == self.relation_dictionary['Equality'] or rel_type == self.relation_dictionary['Equal To']:
                                rep_val = x_repaired[pop_indx, var_index[row_index]]
                                if xl[var_index[col_index]] <= rep_val <= xu[var_index[col_index]]:
                                    x_repaired[pop_indx, var_index[col_index]] = rep_val
                                    var_repaired[pop_indx, var_index[col_index]] = 1
                                else:
                                    warnings.warn(f"Equality relation {rel_type} at ({row_index},{col_index}) requested. But repaired value exceeds limits. Skipping.")
                            elif rel_type == self.relation_dictionary['Power Law']:
                                b, c = float(rel_params[0]), float(rel_params[1])
                                rep_val = np.power(c / x_repaired[pop_indx, var_index[row_index]], 1 / b)
                                if xl[var_index[col_index]] <= rep_val <= xu[var_index[col_index]]:
                                    x_repaired[pop_indx, var_index[col_index]] = rep_val
                                    var_repaired[pop_indx, var_index[col_index]] = 1
                                else:
                                    warnings.warn(f"Power law relation {rel_type} with params {rel_params} requested. But repaired value exceeds limits. Skipping.")
                            else:
                                warnings.warn(f"Relation type {rel_type} specified at ({row_index},{col_index}) of key {key} not "
                                              f"recognized. Skipping.")

        return x_repaired, var_repaired


if __name__ == '__main__':
    np.random.seed(12345)
    rep = KDORepair(n_var=5)
    rep.relation_matrix[rep.DEFAULT_CLUSTER_KEY]['0'][0] = f'N,5,2,0.97'
    rep.relation_confidence[rep.DEFAULT_CLUSTER_KEY][0, 0] = 0.97

    # rep.relation_matrix[rep.DEFAULT_CLUSTER_KEY]['4'][2] = f'P,0.5,2,0.1'
    rep.relation_matrix[rep.DEFAULT_CLUSTER_KEY]['4'][2] = f'P,0.5,2,0.95'
    rep.relation_confidence[rep.DEFAULT_CLUSTER_KEY][2, 4] = 0.95

    c1 = np.round(np.random.rand(), 2)
    # rep.relation_matrix[rep.DEFAULT_CLUSTER_KEY]['2'][0] = f'I,>,{c1}'
    # rep.relation_matrix[rep.DEFAULT_CLUSTER_KEY]['2'][0] = f'I,>,0.92'
    # rep.relation_confidence[rep.DEFAULT_CLUSTER_KEY][0, 2] = 0.92
    rep.relation_matrix[rep.DEFAULT_CLUSTER_KEY]['2'][1] = f'E,0.92'
    rep.relation_confidence[rep.DEFAULT_CLUSTER_KEY][1, 2] = 0.92

    c1 = np.round(np.random.rand(), 2)
    rep.relation_matrix[rep.DEFAULT_CLUSTER_KEY]['3'][0] = f'E,{0.98}'
    rep.relation_confidence[rep.DEFAULT_CLUSTER_KEY][0, 3] = 0.98

    c1 = np.round(np.random.rand(), 2)
    # rep.relation_matrix[rep.DEFAULT_CLUSTER_KEY]['2'][2] = f'U,5,9,{c1}'
    rep.relation_matrix[rep.DEFAULT_CLUSTER_KEY]['2'][2] = f'U,7,9,{0.93}'
    rep.relation_confidence[rep.DEFAULT_CLUSTER_KEY][2, 2] = 0.93

    rep.relation_matrix[rep.DEFAULT_CLUSTER_KEY]['1'][1] = f'C,8.77,{0.95}'
    rep.relation_confidence[rep.DEFAULT_CLUSTER_KEY][1, 1] = 0.95

    rep.write_relation_matrix_to_excel()
    # rep.get_user_feedback()

    # xl = np.zeros(rep.n_var)
    # xu = 10 * np.ones(rep.n_var)
    x_off = np.array([[1., 3., 10., 3., 9.],
                      [7., 0, 9, 9, 6]])

    class problem:
        def __init__(self):
            self.xl = np.zeros(rep.n_var)
            self.xu = 10 * np.ones(rep.n_var)
    class algo:
        def __init__(self):
            self.problem = problem()
    alg = algo()
    x_rep, v_rep = rep.do(problem=alg.problem, pop=x_off, algorithm=alg)
    print("Previous x = ", x_off)
    print("New x = ", x_rep)
